import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

COLUMN_WIDTH = 20


def style_header(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(fill_type="solid", fgColor="2F5496")
    cell.alignment = Alignment(horizontal="center")


def create_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Personnel"

    # --- Soldiers table (A–F) ---
    ws["A1"] = "Soldiers"
    ws["A1"].font = Font(bold=True, size=13)

    for col, header in enumerate(["Name", "Last Name", "Personal Number", "Phone Number", "Course", "Gender"], start=1):
        style_header(ws.cell(row=2, column=col, value=header))
        ws.column_dimensions[get_column_letter(col)].width = COLUMN_WIDTH

    # --- Commanders table (H–J, gap at col G) ---
    ws["H1"] = "Commanders"
    ws["H1"].font = Font(bold=True, size=13)

    for col, header in enumerate(["Name", "Last Name", "Half Kappa"], start=8):
        style_header(ws.cell(row=2, column=col, value=header))
        ws.column_dimensions[get_column_letter(col)].width = COLUMN_WIDTH

    # --- Hitnasuyot table (L–N, gap at col K) ---
    ws["L1"] = "Hitnasuyot"
    ws["L1"].font = Font(bold=True, size=13)

    for col, header in enumerate(["Hitnasut Name", "Full Soldier Name", "Full Commander Name"], start=12):
        style_header(ws.cell(row=2, column=col, value=header))
        ws.column_dimensions[get_column_letter(col)].width = COLUMN_WIDTH

    # --- Courses table (P–Q, gap at col O) ---
    ws["P1"] = "Courses"
    ws["P1"].font = Font(bold=True, size=13)

    for col, header in enumerate(["Course", "Half Kappa"], start=16):
        style_header(ws.cell(row=2, column=col, value=header))
        ws.column_dimensions[get_column_letter(col)].width = COLUMN_WIDTH

    # --- Rooms table (S–V, gap at col R) ---
    ws["S1"] = "Rooms"
    ws["S1"].font = Font(bold=True, size=13)

    for col, header in enumerate(["Room Number", "Gender", "Capacity", "Room Manager"], start=19):
        style_header(ws.cell(row=2, column=col, value=header))
        ws.column_dimensions[get_column_letter(col)].width = COLUMN_WIDTH

    # --- Computer Users table (X–Y, gap at col W) ---
    ws["X1"] = "Computer Users"
    ws["X1"].font = Font(bold=True, size=13)

    for col, header in enumerate(["Username", "Password"], start=24):
        style_header(ws.cell(row=2, column=col, value=header))
        ws.column_dimensions[get_column_letter(col)].width = COLUMN_WIDTH

    # --- Classrooms table (AA–AB, gap at col Z) ---
    ws["AA1"] = "Classrooms"
    ws["AA1"].font = Font(bold=True, size=13)

    for col, header in enumerate(["Full Commander Name", "Class Name"], start=27):
        style_header(ws.cell(row=2, column=col, value=header))
        ws.column_dimensions[get_column_letter(col)].width = COLUMN_WIDTH

    _style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showFirstColumn=False, showLastColumn=False, showColumnStripes=False)
    for name, ref in [
        ("Soldiers",      "A2:F2"),
        ("Commanders",    "H2:J2"),
        ("Hitnasuyot",    "L2:N2"),
        ("Courses",       "P2:Q2"),
        ("Rooms",         "S2:V2"),
        ("ComputerUsers", "X2:Y2"),
        ("Classrooms",    "AA2:AB2"),
    ]:
        tab = Table(displayName=name, ref=ref)
        tab.tableStyleInfo = _style
        ws.add_table(tab)

    wb.save("personnel.xlsx")
    print("Created personnel.xlsx")


if __name__ == "__main__":
    create_workbook()
